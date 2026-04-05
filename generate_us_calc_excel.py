#!/usr/bin/env python3
"""
generate_us_calc_excel.py
=========================
Produces an Excel workbook (US_FedFunds_ImpliedRates_Calc.xlsx) showing
the complete calculation chain for US Fed Funds implied rates:

  Raw TradingView API response
  → Contract-to-meeting mapping
  → Implied rate = 100 − price
  → Base rate (EFFR from FRED DFF)
  → ± bps change

Sheets
------
  1. Raw - TradingView      Raw scanner API response (symbols + prices)
  2. Raw - FRED EFFR         Raw FRED DFF observations
  3. Contract Mapping        FOMC meeting → ZQ contract (month code logic)
  4. Calculation             Step-by-step with Excel formulas
  5. Summary                 Final implied rates matching WIRP dashboard
"""

import json
import sys
from datetime import datetime, date

try:
    import requests
except ImportError:
    sys.exit("Missing: pip install requests")

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing: pip install openpyxl")

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIG  (mirrors update_data.py)
# ─────────────────────────────────────────────────────────────────────────────
FRED_API_KEY = "d5a8cec257d69a6aee9fa88722467ccf"
TV_URL       = "https://scanner.tradingview.com/futures/scan"
FRED_URL     = ("https://api.stlouisfed.org/fred/series/observations"
                "?series_id=DFF&sort_order=desc&limit=10"
                "&file_type=json&api_key=" + FRED_API_KEY)
FED_MEETINGS_URL = "https://www.federalreserve.gov/monetarypolicy/fomccalendars.htm"

_FF_MONTH = {1:'F',2:'G',3:'H',4:'J',5:'K',6:'M',
             7:'N',8:'Q',9:'U',10:'V',11:'X',12:'Z'}

HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

# ─────────────────────────────────────────────────────────────────────────────
#  STYLES
# ─────────────────────────────────────────────────────────────────────────────
AMBER      = "FFD97706"
AMBER_FILL = "FFFFF3CD"
BLUE_HDR   = "FF1D4ED8"
BLUE_FILL  = "FFdbeafe"
GREEN      = "FF15803D"
GREEN_FILL = "FFdcfce7"
RED_FILL   = "FFfee2e2"
GREY_FILL  = "FFF1F5F9"
WHITE      = "FFFFFFFF"

def hdr_style(wb, hex_bg=BLUE_HDR, hex_fg="FFFFFFFF", bold=True):
    f = Font(bold=bold, color=hex_fg, name="Calibri", size=10)
    fill = PatternFill("solid", fgColor=hex_bg)
    al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return f, fill, al

def apply_hdr(cell, wb, hex_bg=BLUE_HDR, hex_fg="FFFFFFFF"):
    f, fill, al = hdr_style(wb, hex_bg, hex_fg)
    cell.font = f
    cell.fill = fill
    cell.alignment = al

def thin_border():
    s = Side(style="thin", color="FFCBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ─────────────────────────────────────────────────────────────────────────────
#  DATA FETCH
# ─────────────────────────────────────────────────────────────────────────────

def fetch_fomc_meetings():
    """Scrape upcoming FOMC meeting dates from federalreserve.gov."""
    from bs4 import BeautifulSoup
    print("Fetching FOMC meeting dates...")
    r = requests.get(FED_MEETINGS_URL, headers=HTTP_HEADERS, timeout=20)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    meetings = []
    today = date.today()
    for p in soup.find_all("p", class_="fomc-meeting__date"):
        txt = p.get_text(" ", strip=True)
        parent = p.find_parent()
        year_tag = parent.find_previous("h4") or parent.find_previous("h5")
        year = None
        if year_tag:
            try:
                year = int(year_tag.get_text(strip=True).strip()[:4])
            except Exception:
                pass
        if not year:
            year = today.year
        # Parse month/day
        import re
        m = re.search(r'([A-Za-z]+)\s+(\d+)(?:[-–]\d+)?', txt)
        if m:
            try:
                dt = datetime.strptime(f"{m.group(1)} {m.group(2)} {year}", "%B %d %Y")
                if dt.date() >= today:
                    meetings.append(dt.strftime("%d %b %Y"))
                    if len(meetings) == 8:
                        break
            except Exception:
                pass
    if not meetings:
        # Fallback hardcoded
        meetings = ["29 Apr 2026","17 Jun 2026","29 Jul 2026","16 Sep 2026",
                    "28 Oct 2026","09 Dec 2026","27 Jan 2027","17 Mar 2027"]
    return meetings


def build_contract_specs(meetings):
    """Build list of (meeting_str, post_month, post_year, tv_symbol, yf_ticker)."""
    specs = []
    for mtg_str in meetings:
        mtg = datetime.strptime(mtg_str, "%d %b %Y").date()
        post_month = mtg.month + 1
        post_year  = mtg.year
        if post_month > 12:
            post_month = 1
            post_year  += 1
        mc      = _FF_MONTH[post_month]
        tv_sym  = f"CBOT:ZQ{mc}{post_year}"
        yf_tick = f"ZQ{mc}{str(post_year)[-2:]}.CBT"
        specs.append((mtg_str, post_month, post_year, mc, tv_sym, yf_tick))
    return specs


def fetch_tv_prices(tickers):
    """Call TradingView scanner API. Returns (raw_json_str, dict {sym: price})."""
    print(f"Fetching TradingView prices for {len(tickers)} contracts...")
    payload = {"symbols": {"tickers": tickers}, "columns": ["close", "description"]}
    r = requests.post(TV_URL, json=payload, headers={
        "User-Agent": HTTP_HEADERS["User-Agent"],
        "Content-Type": "application/json",
    }, timeout=20)
    r.raise_for_status()
    data = r.json()
    raw = json.dumps(data, indent=2)
    prices = {}
    for item in data.get("data", []):
        sym  = item["s"]
        vals = item.get("d", [])
        if vals and vals[0] is not None:
            prices[sym] = float(vals[0])
    return raw, data, prices


def fetch_fred_effr():
    """Fetch last 10 DFF observations from FRED. Returns (raw_json_str, list of obs, latest_rate)."""
    print("Fetching FRED DFF (EFFR)...")
    r = requests.get(FRED_URL, timeout=20)
    r.raise_for_status()
    data = r.json()
    raw  = json.dumps(data, indent=2)
    obs  = data.get("observations", [])
    latest = None
    for o in obs:
        v = o.get("value", ".")
        if v != ".":
            latest = float(v)
            break
    return raw, obs, latest


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(meetings, specs, tv_raw, tv_data, tv_prices, fred_raw, fred_obs, effr):
    wb = openpyxl.Workbook()

    # ── SHEET 1: Raw TradingView ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "1. Raw - TradingView"
    _sheet_tv(ws1, tv_raw, tv_data, specs)

    # ── SHEET 2: Raw FRED EFFR ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("2. Raw - FRED EFFR")
    _sheet_fred(ws2, fred_raw, fred_obs)

    # ── SHEET 3: Contract Mapping ─────────────────────────────────────────────
    ws3 = wb.create_sheet("3. Contract Mapping")
    _sheet_mapping(ws3, specs, tv_prices)

    # ── SHEET 4: Calculation ──────────────────────────────────────────────────
    ws4 = wb.create_sheet("4. Calculation")
    _sheet_calc(ws4, specs, tv_prices, effr)

    # ── SHEET 5: Summary ──────────────────────────────────────────────────────
    ws5 = wb.create_sheet("5. Summary")
    _sheet_summary(ws5, specs, tv_prices, effr)

    return wb


def _title_row(ws, text, row=1):
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=13, color=AMBER, name="Calibri")
    c.fill = PatternFill("solid", fgColor="FF1E293B")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

def _meta_row(ws, label, value, row):
    ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=9, color="FF64748B", name="Calibri")
    c = ws.cell(row=row, column=2, value=value)
    c.font = Font(size=9, name="Calibri")
    c.alignment = Alignment(wrap_text=True)

def _hdr(ws, row, cols_labels, bg=BLUE_HDR):
    for col, label in enumerate(cols_labels, 1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = Font(bold=True, size=9, color="FFFFFFFF", name="Calibri")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 30

def _cell(ws, row, col, value, bold=False, align="left", fmt=None, bg=None, color="FF1E293B"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=9, name="Calibri", color=color)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = thin_border()
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c


def _sheet_tv(ws, raw_str, tv_data, specs):
    _title_row(ws, "Raw TradingView Scanner API Response", 1)
    _meta_row(ws, "Endpoint:", TV_URL, 2)
    _meta_row(ws, "Method:", "POST", 3)
    _meta_row(ws, "Timestamp:", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"), 4)
    _meta_row(ws, "Payload:", '{"symbols":{"tickers":[...]},"columns":["close","description"]}', 5)

    # Request payload section
    ws.cell(row=7, column=1, value="REQUEST PAYLOAD (sent to API)").font = Font(bold=True, size=10, name="Calibri", color="FF1D4ED8")
    payload_tickers = [s[4] for s in specs]
    payload = {"symbols": {"tickers": payload_tickers}, "columns": ["close", "description"]}
    payload_str = json.dumps(payload, indent=2)
    ws.cell(row=8, column=1, value=payload_str).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A8:H8")
    ws.row_dimensions[8].height = max(60, len(payload_str.splitlines()) * 12)

    # Response section
    ws.cell(row=10, column=1, value="RESPONSE — data array").font = Font(bold=True, size=10, name="Calibri", color="FF1D4ED8")
    _hdr(ws, 11, ["#", "Symbol (s)", "Close Price (d[0])", "Description (d[1])", "Source", "Status"])

    row = 12
    for i, item in enumerate(tv_data.get("data", []), 1):
        sym  = item.get("s", "")
        vals = item.get("d", [])
        price = vals[0] if vals and vals[0] is not None else "N/A"
        desc  = vals[1] if len(vals) > 1 and vals[1] is not None else ""
        bg = WHITE if i % 2 else GREY_FILL
        _cell(ws, row, 1, i,     align="center", bg=bg)
        _cell(ws, row, 2, sym,   bold=True, bg=bg)
        _cell(ws, row, 3, price, align="right", fmt="0.0000", bg=bg,
              color=GREEN if price != "N/A" else "FFDC2626", bold=price != "N/A")
        _cell(ws, row, 4, desc,  bg=bg)
        _cell(ws, row, 5, "TradingView Scanner", bg=bg)
        _cell(ws, row, 6, "✓ Received" if price != "N/A" else "✗ Missing",
              bg=bg, color=GREEN if price != "N/A" else "FFDC2626")
        row += 1

    # Raw JSON
    ws.cell(row=row+1, column=1, value="FULL RAW JSON RESPONSE").font = Font(bold=True, size=10, name="Calibri", color="FF1D4ED8")
    c = ws.cell(row=row+2, column=1, value=raw_str)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(name="Courier New", size=8, color="FF334155")
    ws.merge_cells(f"A{row+2}:H{row+2}")
    ws.row_dimensions[row+2].height = max(80, min(400, len(raw_str.splitlines()) * 11))

    set_col_width(ws, 1, 6)
    set_col_width(ws, 2, 22)
    set_col_width(ws, 3, 18)
    set_col_width(ws, 4, 40)
    set_col_width(ws, 5, 24)
    set_col_width(ws, 6, 16)
    ws.freeze_panes = "A12"


def _sheet_fred(ws, raw_str, obs):
    _title_row(ws, "Raw FRED API Response — DFF (Effective Federal Funds Rate)", 1)
    _meta_row(ws, "Endpoint:", FRED_URL.split("?")[0], 2)
    _meta_row(ws, "Series:", "DFF — Effective Federal Funds Rate (daily)", 3)
    _meta_row(ws, "Timestamp:", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"), 4)
    _meta_row(ws, "Note:", "WIRP uses the most recent non-'.' observation as the base rate (EFFR)", 5)

    ws.cell(row=7, column=1, value="OBSERVATIONS (last 10, newest first)").font = Font(bold=True, size=10, name="Calibri", color="FF1D4ED8")
    _hdr(ws, 8, ["#", "Date", "Value (%)", "Used as Base Rate?", "Notes"])

    latest_used = False
    for i, o in enumerate(obs[:10], 1):
        v   = o.get("value", ".")
        dt  = o.get("date", "")
        is_latest = (not latest_used and v != ".")
        if is_latest:
            latest_used = True
        bg = AMBER_FILL if is_latest else (WHITE if i % 2 else GREY_FILL)
        _cell(ws, 8+i, 1, i, align="center", bg=bg)
        _cell(ws, 8+i, 2, dt, bg=bg)
        _cell(ws, 8+i, 3, float(v) if v != "." else ".", align="right", fmt="0.0000", bg=bg,
              bold=is_latest)
        _cell(ws, 8+i, 4, "✓ YES — used as EFFR base rate" if is_latest else "No",
              bg=bg, color="FFd97706" if is_latest else "FF64748B", bold=is_latest)
        _cell(ws, 8+i, 5, "Most recent non-missing value" if is_latest else
              ("Value='.' = not yet published" if v == "." else ""), bg=bg)

    row = 19
    ws.cell(row=row, column=1, value="FULL RAW JSON RESPONSE").font = Font(bold=True, size=10, name="Calibri", color="FF1D4ED8")
    c = ws.cell(row=row+1, column=1, value=raw_str)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(name="Courier New", size=8, color="FF334155")
    ws.merge_cells(f"A{row+1}:H{row+1}")
    ws.row_dimensions[row+1].height = max(80, min(300, len(raw_str.splitlines()) * 11))

    set_col_width(ws, 1, 5)
    set_col_width(ws, 2, 14)
    set_col_width(ws, 3, 14)
    set_col_width(ws, 4, 32)
    set_col_width(ws, 5, 40)
    ws.freeze_panes = "A9"


def _sheet_mapping(ws, specs, tv_prices):
    _title_row(ws, "FOMC Meeting → ZQ Contract Mapping", 1)

    ws.cell(row=3, column=1, value="METHODOLOGY").font = Font(bold=True, size=10, name="Calibri", color="FF1D4ED8")
    method_text = (
        "For an FOMC meeting in month M of year Y:\n"
        "  → Use the CME 30-Day Fed Funds futures contract for month M+1\n"
        "     (the post-meeting month captures the new rate for the full month)\n\n"
        "Contract symbol format:  CBOT:ZQ{month_code}{year}\n"
        "Month codes: Jan=F  Feb=G  Mar=H  Apr=J  May=K  Jun=M\n"
        "             Jul=N  Aug=Q  Sep=U  Oct=V  Nov=X  Dec=Z\n\n"
        "Example:  FOMC on 17 Jun 2026 (month 6)\n"
        "          Post-meeting month = 7 (Jul) → month code = N\n"
        "          Contract = CBOT:ZQN2026\n\n"
        "Rationale: The 30-Day Fed Funds futures for month M+1 settle at the\n"
        "average effective federal funds rate for that month.  Since the FOMC\n"
        "decision takes effect immediately after the meeting, the M+1 contract\n"
        "captures the new rate for its full settlement month."
    )
    c = ws.cell(row=4, column=1, value=method_text)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(size=9, name="Calibri", color="FF334155")
    ws.merge_cells("A4:H4")
    ws.row_dimensions[4].height = 160

    _hdr(ws, 6, [
        "FOMC Meeting Date",
        "Meeting Month (M)",
        "Post-Meeting Month (M+1)",
        "Month Code",
        "Contract Year",
        "TradingView Symbol",
        "Yahoo Finance Ticker",
        "Price Available?",
    ])

    for i, (mtg_str, post_month, post_year, mc, tv_sym, yf_tick) in enumerate(specs, 1):
        mtg_dt = datetime.strptime(mtg_str, "%d %b %Y")
        bg = WHITE if i % 2 else GREY_FILL
        has_price = tv_sym in tv_prices
        _cell(ws, 6+i, 1, mtg_str, bold=True, bg=bg)
        _cell(ws, 6+i, 2, f"{mtg_dt.month} ({mtg_dt.strftime('%b')})", align="center", bg=bg)
        _cell(ws, 6+i, 3, f"{post_month} ({datetime(2000,post_month,1).strftime('%b')})", align="center", bg=bg)
        _cell(ws, 6+i, 4, mc, align="center", bold=True, bg=bg, color="FF1D4ED8")
        _cell(ws, 6+i, 5, post_year, align="center", bg=bg)
        _cell(ws, 6+i, 6, tv_sym, bold=True, bg=BLUE_FILL if has_price else RED_FILL,
              color="FF1D4ED8" if has_price else "FFDC2626")
        _cell(ws, 6+i, 7, yf_tick, bg=bg)
        _cell(ws, 6+i, 8,
              "✓ Received" if has_price else "✗ Missing",
              bg=GREEN_FILL if has_price else RED_FILL,
              color=GREEN if has_price else "FFDC2626",
              bold=True, align="center")

    set_col_width(ws, 1, 20)
    set_col_width(ws, 2, 18)
    set_col_width(ws, 3, 22)
    set_col_width(ws, 4, 14)
    set_col_width(ws, 5, 14)
    set_col_width(ws, 6, 20)
    set_col_width(ws, 7, 20)
    set_col_width(ws, 8, 16)
    ws.freeze_panes = "A7"


def _sheet_calc(ws, specs, tv_prices, effr):
    _title_row(ws, "Step-by-Step Calculation — US Fed Funds Implied Rates", 1)

    # EFFR box
    ws.cell(row=3, column=1, value="BASE RATE INPUT").font = Font(bold=True, size=10, name="Calibri", color="FF1D4ED8")
    _hdr(ws, 4, ["Source", "Series", "Latest Date", "EFFR Value (%)"], bg="FF1D4ED8")
    _cell(ws, 5, 1, "FRED API", bold=True, bg=AMBER_FILL)
    _cell(ws, 5, 2, "DFF — Effective Federal Funds Rate", bg=AMBER_FILL)
    _cell(ws, 5, 3, datetime.utcnow().strftime("%Y-%m-%d"), bg=AMBER_FILL)
    c = ws.cell(row=5, column=4, value=effr)
    c.font = Font(bold=True, size=12, name="Calibri", color=AMBER)
    c.fill = PatternFill("solid", fgColor=AMBER_FILL)
    c.number_format = "0.0000"
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border()
    # Name the EFFR cell for formula references
    ws.cell(row=5, column=4).value = effr

    ws.row_dimensions[5].height = 18

    # Calculation table
    ws.cell(row=7, column=1, value="CALCULATION TABLE").font = Font(bold=True, size=10, name="Calibri", color="FF1D4ED8")

    hdrs = [
        "FOMC Meeting",
        "ZQ Contract",
        "TradingView\nClose Price",
        "Step 1\nImplied Rate (%)\n= 100 − Price",
        "Step 2\nBase Rate\n(EFFR from FRED)",
        "Step 3\nChange (%)\n= Implied − Base",
        "Step 4\n± bps\n= Change × 100",
        "Direction",
    ]
    _hdr(ws, 8, hdrs, bg="FF1E293B")

    for i, (mtg_str, post_month, post_year, mc, tv_sym, yf_tick) in enumerate(specs, 1):
        row = 8 + i
        price = tv_prices.get(tv_sym)
        has   = price is not None
        bg    = WHITE if i % 2 else GREY_FILL

        implied = round(100.0 - price, 4) if has else None
        change  = round(implied - effr, 4) if (has and effr is not None) else None
        bps     = round(change * 100, 4)   if change is not None else None

        _cell(ws, row, 1, mtg_str, bold=True, bg=bg)
        _cell(ws, row, 2, tv_sym, bg=bg, color="FF1D4ED8")
        if has:
            _cell(ws, row, 3, price,   align="right", fmt="0.0000", bg=bg)
            # Use an Excel formula for transparency
            pc = get_column_letter(3)
            _cell(ws, row, 4, implied, align="right", fmt="0.0000",
                  bg=BLUE_FILL, bold=True)
            ws.cell(row=row, column=4).value = f"=100-{pc}{row}"

            ec = "D5"   # EFFR cell
            _cell(ws, row, 5, effr,   align="right", fmt="0.0000", bg=AMBER_FILL)
            ws.cell(row=row, column=5).value = f"=$D$5"

            _cell(ws, row, 6, change,  align="right", fmt="0.00000",
                  bg=GREEN_FILL if (change or 0) >= 0 else RED_FILL,
                  color=GREEN if (change or 0) >= 0 else "FFDC2626")
            ws.cell(row=row, column=6).value = f"=D{row}-E{row}"

            _cell(ws, row, 7, bps, align="right", fmt="0.0000",
                  bg=GREEN_FILL if (bps or 0) >= 0 else RED_FILL,
                  color=GREEN if (bps or 0) >= 0 else "FFDC2626",
                  bold=True)
            ws.cell(row=row, column=7).value = f"=F{row}*100"

            direction = ("▲ Hike" if (bps or 0) >= 1 else
                         "▼ Cut"  if (bps or 0) <= -1 else "━ Hold")
            dir_color = (GREEN if (bps or 0) >= 1 else
                         "FFDC2626" if (bps or 0) <= -1 else AMBER)
            _cell(ws, row, 8, direction, align="center", bold=True,
                  bg=bg, color=dir_color)
        else:
            for col in range(3, 9):
                _cell(ws, row, col, "No price data", align="center",
                      bg=RED_FILL, color="FFDC2626")

    # Formula legend
    last_data_row = 8 + len(specs)
    leg_row = last_data_row + 2
    ws.cell(row=leg_row, column=1, value="FORMULA REFERENCE").font = Font(bold=True, size=10, name="Calibri", color="FF1D4ED8")
    formulas = [
        ("Step 1 — Implied Rate",  "= 100 − Futures Price",
         "ZQ futures are quoted as 100 minus the expected average overnight rate. "
         "A price of 96.365 implies a rate of 3.635%."),
        ("Step 2 — Base Rate",     "= EFFR (latest FRED DFF)",
         "The Effective Federal Funds Rate published daily by the New York Fed, "
         "sourced via FRED API. This is the current realised rate, matching Bloomberg WIRP."),
        ("Step 3 — Change",        "= Implied Rate − Base Rate",
         "Positive = market pricing a hike; Negative = cut priced in."),
        ("Step 4 — bps",           "= Change × 100",
         "Converts percentage-point change to basis points (1% = 100 bps). "
         "Granularity is ~0.5 bps — the minimum tick of ZQ futures (0.005 price increment)."),
    ]
    _hdr(ws, leg_row+1, ["Formula", "Expression", "Explanation"], bg="FF334155")
    for j, (name, expr, explain) in enumerate(formulas, 1):
        bg = WHITE if j % 2 else GREY_FILL
        _cell(ws, leg_row+1+j, 1, name,    bold=True, bg=bg, color="FF1D4ED8")
        _cell(ws, leg_row+1+j, 2, expr,    bold=True, bg=BLUE_FILL)
        _cell(ws, leg_row+1+j, 3, explain, bg=bg)
        ws.merge_cells(f"C{leg_row+1+j}:H{leg_row+1+j}")

    set_col_width(ws, 1, 20)
    set_col_width(ws, 2, 20)
    set_col_width(ws, 3, 16)
    set_col_width(ws, 4, 18)
    set_col_width(ws, 5, 16)
    set_col_width(ws, 6, 16)
    set_col_width(ws, 7, 14)
    set_col_width(ws, 8, 12)
    ws.freeze_panes = "A9"
    ws.row_dimensions[8].height = 42


def _sheet_summary(ws, specs, tv_prices, effr):
    _title_row(ws, "Summary — US Fed Funds Implied Rates (WIRP Output)", 1)
    _meta_row(ws, "Generated:", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"), 2)
    _meta_row(ws, "Base Rate (EFFR):", f"{effr:.4f}%" if effr else "N/A", 3)
    _meta_row(ws, "Source:", "CME ZQ 30-Day Fed Funds Futures via TradingView Scanner API", 4)

    _hdr(ws, 6, [
        "#", "FOMC Meeting", "ZQ Contract", "Futures Price",
        "Implied Rate (%)", "Base Rate (%)", "± bps", "Direction",
    ])

    for i, (mtg_str, post_month, post_year, mc, tv_sym, yf_tick) in enumerate(specs, 1):
        price   = tv_prices.get(tv_sym)
        implied = round(100.0 - price, 4) if price is not None else None
        change  = round(implied - effr, 4)   if (implied and effr) else None
        bps     = round(change * 100, 4)     if change is not None else None
        bg      = WHITE if i % 2 else GREY_FILL

        direction = ("▲ Hike" if (bps or 0) >= 1 else
                     "▼ Cut"  if (bps or 0) <= -1 else "━ Hold") if bps is not None else "—"
        dir_color = (GREEN if (bps or 0) >= 1 else
                     "FFDC2626" if (bps or 0) <= -1 else AMBER) if bps is not None else "FF64748B"

        _cell(ws, 6+i, 1, i,         align="center", bg=bg)
        _cell(ws, 6+i, 2, mtg_str,   bold=True, bg=bg)
        _cell(ws, 6+i, 3, tv_sym,    bg=BLUE_FILL if price else RED_FILL,
              color="FF1D4ED8" if price else "FFDC2626")
        _cell(ws, 6+i, 4, price,     align="right", fmt="0.0000", bg=bg)
        _cell(ws, 6+i, 5, implied,   align="right", fmt="0.0000",
              bg=BLUE_FILL, bold=True, color="FF1D4ED8")
        _cell(ws, 6+i, 6, effr,      align="right", fmt="0.0000", bg=AMBER_FILL)
        _cell(ws, 6+i, 7, bps,       align="right", fmt="0.0000",
              bg=GREEN_FILL if (bps or 0) >= 0 else RED_FILL,
              color=GREEN if (bps or 0) >= 0 else "FFDC2626", bold=True)
        _cell(ws, 6+i, 8, direction, align="center", bold=True, bg=bg, color=dir_color)

    set_col_width(ws, 1, 5)
    set_col_width(ws, 2, 20)
    set_col_width(ws, 3, 20)
    set_col_width(ws, 4, 15)
    set_col_width(ws, 5, 16)
    set_col_width(ws, 6, 14)
    set_col_width(ws, 7, 12)
    set_col_width(ws, 8, 12)
    ws.freeze_panes = "A7"


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import os
    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "US_FedFunds_ImpliedRates_Calc.xlsx")

    try:
        meetings = fetch_fomc_meetings()
    except Exception as e:
        print(f"Warning: could not scrape Fed meetings ({e}), using hardcoded fallback")
        meetings = ["29 Apr 2026","17 Jun 2026","29 Jul 2026","16 Sep 2026",
                    "28 Oct 2026","09 Dec 2026","27 Jan 2027","17 Mar 2027"]

    print(f"Meetings: {meetings}")
    specs = build_contract_specs(meetings)

    tv_tickers = [s[4] for s in specs]
    tv_raw, tv_data, tv_prices = fetch_tv_prices(tv_tickers)
    print(f"TradingView prices received: {tv_prices}")

    fred_raw, fred_obs, effr = fetch_fred_effr()
    print(f"EFFR (base rate): {effr}%")

    print("Building Excel workbook...")
    wb = build_excel(meetings, specs, tv_raw, tv_data, tv_prices,
                     fred_raw, fred_obs, effr)

    wb.save(out_path)
    print(f"\nSaved: {out_path}")
