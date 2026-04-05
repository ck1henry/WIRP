#!/usr/bin/env python3
"""
WIRP Data Updater — G10 + Key Asian Economies
==============================================
Scrapes live central bank meeting dates and current policy rates from
official sources, then injects the data directly into wirp.html.

Markets covered
---------------
  AMERICAS : US (Fed), CA (BOC)
  EUROPE   : EU (ECB), UK (BOE)
  ASIA-PAC : AU (RBA)

All five markets use live scrapers for meeting dates and policy rates.
Implied rates come from real OIS / exchange-traded futures where available.

Usage
-----
  python update_data.py            # run once
  python update_data.py --dry-run  # print without writing

Windows Task Scheduler setup — see bottom of file.

Dependencies
------------
  pip install requests beautifulsoup4
"""

import argparse
import json
import logging
import os
import re
import sys
from datetime import datetime, date, timezone

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    print("Missing dependencies.  Run:  pip install requests beautifulsoup4")
    sys.exit(1)

try:
    import openpyxl
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
HTML_FILE  = os.path.join(SCRIPT_DIR, "wirp.html")
JSON_FILE  = os.path.join(SCRIPT_DIR, "wirp_data.json")
LOG_FILE   = os.path.join(SCRIPT_DIR, "wirp_update.log")
HISTORY_FILE = os.path.join(SCRIPT_DIR, "wirp_history.json")
HISTORY_MAX_DAYS = 1095  # 3 years

DATA_BEGIN = "<!-- WIRP_DATA_BEGIN -->"
DATA_END   = "<!-- WIRP_DATA_END -->"

N_MEETINGS = 8    # target number of upcoming meetings to display

# ── FRED API key — used to fetch the daily Effective Fed Funds Rate (DFF) ─────
# Free key at https://fred.stlouisfed.org/docs/api/api_key.html
FRED_API_KEY = "d5a8cec257d69a6aee9fa88722467ccf"

# ── Drift — total implied rate change (%) across next N meetings ─────────────
# Negative = net cuts priced; Positive = net hikes.
# Calibrate against OIS/futures when market positioning shifts materially.
DRIFT = {
    # Fallback only — overridden by real OIS/futures data when available
    "US": 0.00,
    "EU": 0.00,
    "UK": 0.00,
    "CA": 0.00,
    "AU": 0.00,
}

# ── Static metadata ──────────────────────────────────────────────────────────
# live=True  → meetings + rate scraped from official website each run
# live=False → hardcoded fallback; update FALLBACK manually after decisions
META = {
    "US": {"name": "Federal Reserve",           "abbr": "FOMC",   "region": "Americas", "live": True, "source": "CME ZQ futures (TradingView / Yahoo Finance)"},
    "EU": {"name": "European Central Bank",     "abbr": "ECB GC", "region": "Europe",   "live": True, "source": "Eurex EURIBOR 3M futures (TradingView)"},
    "UK": {"name": "Bank of England",           "abbr": "MPC",    "region": "Europe",   "live": True, "source": "ICE 3M SONIA futures (TradingView)"},
    "CA": {"name": "Bank of Canada",            "abbr": "BOC",    "region": "Americas", "live": True, "source": "TMX 3M CORRA futures (TradingView)"},
    "AU": {"name": "Reserve Bank of Australia", "abbr": "RBA",    "region": "Asia-Pac", "live": True, "source": "ASX 30-day IB cash rate futures (TradingView)"},
}

# ── Hardcoded fallback data ───────────────────────────────────────────────────
# Used when scraping fails or is not attempted (JS-rendered / blocked sites).
# Rates and dates valid as of April 2026 — update after each policy decision.
FALLBACK = {
    "US": {
        "rate": 3.625,
        "meetings": ["29 Apr 2026","17 Jun 2026","29 Jul 2026","16 Sep 2026",
                     "28 Oct 2026","09 Dec 2026","27 Jan 2027","17 Mar 2027"],
    },
    "EU": {
        "rate": 2.00,
        "meetings": ["08 Apr 2026","30 Apr 2026","11 Jun 2026","23 Jul 2026",
                     "10 Sep 2026","29 Oct 2026","17 Dec 2026","11 Mar 2027"],
    },
    "UK": {
        "rate": 4.25,
        "meetings": ["08 May 2026","18 Jun 2026","06 Aug 2026","17 Sep 2026",
                     "05 Nov 2026","18 Dec 2026","05 Feb 2027","18 Mar 2027"],
    },
    "CA": {
        "rate": 2.75,
        "meetings": ["16 Apr 2026","04 Jun 2026","15 Jul 2026","09 Sep 2026",
                     "21 Oct 2026","09 Dec 2026","21 Jan 2027","04 Mar 2027"],
    },
    "AU": {
        "rate": 4.10,
        "meetings": ["19 May 2026","07 Jul 2026","11 Aug 2026","22 Sep 2026",
                     "03 Nov 2026","08 Dec 2026","03 Feb 2027","17 Mar 2027"],
    },
}

HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}

# ─────────────────────────────────────────────────────────────────────────────
#  LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(
            open(sys.stdout.fileno(), mode="w", encoding="utf-8", closefd=False)
        ),
    ],
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
#  HTTP HELPER
# ─────────────────────────────────────────────────────────────────────────────
def fetch(url: str, timeout: int = 20) -> BeautifulSoup:
    resp = requests.get(url, headers=HTTP_HEADERS, timeout=timeout)
    resp.raise_for_status()
    return BeautifulSoup(resp.content, "html.parser")


# ─────────────────────────────────────────────────────────────────────────────
#  ═══════════  MEETING DATE SCRAPERS  ═══════════
# ─────────────────────────────────────────────────────────────────────────────

def scrape_fed_meetings() -> list[str]:
    """
    FOMC calendars page.
      .fomc-meeting__month  → "January"
      .fomc-meeting__date   → "27-28"   (decision on last day)
    Year derived from PDF link: monetary{YYYYMMDD}...
    """
    url  = "https://www.federalreserve.gov/monetarypolicy/fomccalendars.htm"
    log.info("Fetching US meetings: %s", url)
    soup  = fetch(url)
    today = date.today()
    meetings: list[str] = []

    for row in soup.find_all("div", class_="fomc-meeting"):
        month_el = row.find(class_=re.compile(r"fomc-meeting__month"))
        date_el  = row.find(class_=re.compile(r"fomc-meeting__date"))
        if not (month_el and date_el):
            continue
        try:
            month_num = datetime.strptime(month_el.get_text(strip=True).strip()[:3], "%b").month
        except ValueError:
            continue
        days_raw = re.sub(r"[^\d\-]", "", date_el.get_text(strip=True))
        if not days_raw:
            continue
        day = int(days_raw.split("-")[-1]) if "-" in days_raw else int(days_raw)

        year: int | None = None
        for a in row.find_all("a", href=True):
            m = re.search(r"monetary(\d{4})\d{4}", a["href"])
            if m: year = int(m.group(1)); break
        if year is None:
            for a in row.find_all("a", href=True):
                m = re.search(r"/(\d{4})/", a["href"])
                if m: year = int(m.group(1)); break
        if year is None:
            for y in (today.year, today.year + 1):
                try:
                    if date(y, month_num, day) >= today: year = y; break
                except ValueError: pass
        if year is None: continue
        try:
            dt = date(year, month_num, day)
        except ValueError: continue
        if dt >= today:
            meetings.append(dt.strftime("%d %b %Y"))
        if len(meetings) == N_MEETINGS: break

    log.info("  -> %d US meetings", len(meetings))
    return meetings


def scrape_boe_meetings() -> list[str]:
    """
    BOE upcoming MPC dates page.
    First <td> per row: "Thursday 18 June"
    Year from href: /YYYY/month-YYYY
    """
    url  = "https://www.bankofengland.co.uk/monetary-policy/upcoming-mpc-dates"
    log.info("Fetching UK meetings: %s", url)
    soup  = fetch(url)
    today = date.today()
    meetings: list[str] = []

    for tr in soup.find_all("tr"):
        tds = tr.find_all("td")
        if not tds: continue
        m = re.search(r"(\d{1,2})\s+([A-Za-z]+)", tds[0].get_text(strip=True))
        if not m: continue
        try:
            month_num = datetime.strptime(m.group(2)[:3], "%b").month
        except ValueError: continue
        day = int(m.group(1))
        year: int | None = None
        for a in tr.find_all("a", href=True):
            ym = re.search(r"/(\d{4})/", a["href"])
            if ym: year = int(ym.group(1)); break
        if year is None:
            for y in (today.year, today.year + 1):
                try:
                    if date(y, month_num, day) >= today: year = y; break
                except ValueError: pass
        if year is None: continue
        try:
            dt = date(year, month_num, day)
        except ValueError: continue
        if dt >= today:
            meetings.append(dt.strftime("%d %b %Y"))
        if len(meetings) == N_MEETINGS: break

    log.info("  -> %d UK meetings", len(meetings))
    return meetings


def scrape_ecb_meetings() -> list[str]:
    """
    ECB GC calendar.
    <dt>DD/MM/YYYY</dt> paired with <dd> description.
    Filter for Day 2 / press conference = decision day.
    """
    url  = "https://www.ecb.europa.eu/press/calendars/mgcgc/html/index.en.html"
    log.info("Fetching EU meetings: %s", url)
    soup  = fetch(url)
    today = date.today()
    meetings: list[str] = []

    dl = soup.find("div", class_="definition-list") or soup.find("dl")
    if not dl:
        raise ValueError("ECB calendar <dl> not found")

    for dt_el, dd_el in zip(dl.find_all("dt"), dl.find_all("dd")):
        dm = re.match(r"(\d{2})/(\d{2})/(\d{4})", dt_el.get_text(strip=True))
        if not dm: continue
        desc = dd_el.get_text(strip=True).lower()
        if "day 2" not in desc and "press conference" not in desc: continue
        try:
            dt = date(int(dm.group(3)), int(dm.group(2)), int(dm.group(1)))
        except ValueError: continue
        if dt >= today:
            meetings.append(dt.strftime("%d %b %Y"))
        if len(meetings) == N_MEETINGS: break

    log.info("  -> %d EU meetings", len(meetings))
    return meetings


def scrape_boc_meetings() -> list[str]:
    """
    Bank of Canada key-interest-rate page.
    Table #target-table contains historical dates with data-date="YYYY-MM-DD".
    We pick the most recent N_MEETINGS entries that are today or later.
    If insufficient future dates exist in the table, fall back to FALLBACK data.
    """
    url  = "https://www.bankofcanada.ca/core-functions/monetary-policy/key-interest-rate/"
    log.info("Fetching CA meetings: %s", url)
    soup  = fetch(url)
    today = date.today()
    meetings: list[str] = []

    table = soup.find("table", id="target-table")
    if not table:
        raise ValueError("BOC #target-table not found")

    for th in table.find_all("th", attrs={"data-date": True}):
        ds = th["data-date"]
        try:
            dt = datetime.strptime(ds, "%Y-%m-%d").date()
        except ValueError: continue
        if dt >= today:
            meetings.append(dt.strftime("%d %b %Y"))

    # The table is historical; upcoming dates may be sparse — supplement with fallback
    if len(meetings) < N_MEETINGS:
        fallback_dates = FALLBACK["CA"]["meetings"]
        existing = set(meetings)
        for fd in fallback_dates:
            try:
                dt = datetime.strptime(fd, "%d %b %Y").date()
            except ValueError: continue
            if dt >= today and fd not in existing:
                meetings.append(fd)
            if len(meetings) == N_MEETINGS: break
        meetings.sort(key=lambda s: datetime.strptime(s, "%d %b %Y"))

    log.info("  -> %d CA meetings", len(meetings))
    return meetings[:N_MEETINGS]


def scrape_rba_meetings() -> list[str]:
    """
    RBA interest rate decisions page.
    <ul><li><a>DD Month YYYY</a></li></ul>
    Filters for today-and-future dates.
    """
    url  = "https://www.rba.gov.au/monetary-policy/int-rate-decisions/"
    log.info("Fetching AU meetings: %s", url)
    soup  = fetch(url)
    today = date.today()
    meetings: list[str] = []

    for a in soup.find_all("a", href=True):
        text = a.get_text(strip=True)
        m = re.match(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", text)
        if not m: continue
        try:
            dt = datetime.strptime(f"{m.group(1)} {m.group(2)[:3]} {m.group(3)}", "%d %b %Y").date()
        except ValueError: continue
        if dt >= today:
            meetings.append(dt.strftime("%d %b %Y"))
        if len(meetings) == N_MEETINGS: break

    # Supplement with fallback if page only shows past decisions
    if len(meetings) < 4:
        log.warning("  RBA page returned few future dates; supplementing from fallback")
        fallback_dates = FALLBACK["AU"]["meetings"]
        existing = set(meetings)
        for fd in fallback_dates:
            try:
                dt = datetime.strptime(fd, "%d %b %Y").date()
            except ValueError: continue
            if dt >= today and fd not in existing:
                meetings.append(fd)
            if len(meetings) == N_MEETINGS: break
        meetings.sort(key=lambda s: datetime.strptime(s, "%d %b %Y"))

    log.info("  -> %d AU meetings", len(meetings))
    return meetings[:N_MEETINGS]


# ─────────────────────────────────────────────────────────────────────────────
#  ═══════════  RATE SCRAPERS  ═══════════
# ─────────────────────────────────────────────────────────────────────────────

def scrape_fed_rate() -> float:
    """
    Primary: FRED API series DFF (daily Effective Federal Funds Rate).
    Fallback: Fed target range midpoint scraped from federalreserve.gov.
    Bloomberg WIRP uses the effective rate, not the midpoint, so this
    ensures our ±bps figures are on the same basis.
    """
    if FRED_API_KEY:
        try:
            url = (
                "https://api.stlouisfed.org/fred/series/observations"
                f"?series_id=DFF&sort_order=desc&limit=5"
                f"&file_type=json&api_key={FRED_API_KEY}"
            )
            r = requests.get(url, timeout=15)
            r.raise_for_status()
            for obs in r.json().get("observations", []):
                v = obs.get("value", ".")
                if v != ".":
                    rate = round(float(v), 4)
                    log.info("  -> US rate (EFFR/DFF): %.4f%%", rate)
                    return rate
        except Exception as exc:
            log.warning("FRED DFF failed: %s — falling back to Fed website", exc)

    # Fallback: scrape target range midpoint from Fed website
    url = "https://www.federalreserve.gov/monetarypolicy/openmarket.htm"
    log.info("Fetching US rate: %s", url)
    soup = fetch(url)
    tbody = soup.find("tbody")
    if tbody:
        row = tbody.find("tr")
        if row:
            tds = row.find_all("td")
            if tds:
                m = re.match(r"([\d.]+)[-–]([\d.]+)", tds[-1].get_text(strip=True))
                if m:
                    rate = round((float(m.group(1)) + float(m.group(2))) / 2, 4)
                    log.info("  -> US rate (midpoint fallback): %.4f%%", rate)
                    return rate
    raise ValueError("US rate not found")


def scrape_boe_rate() -> float:
    url = "https://www.bankofengland.co.uk/monetary-policy/the-interest-rate-bank-rate"
    log.info("Fetching UK rate: %s", url)
    soup = fetch(url)
    el = soup.find(class_="stat-figure")
    if el:
        m = re.search(r"([\d.]+)", el.get_text())
        if m:
            rate = float(m.group(1))
            log.info("  -> UK rate: %.4f%%", rate)
            return rate
    raise ValueError("UK rate not found")


def scrape_ecb_rate() -> float:
    """Primary: ECB SDW REST API. Fallback: HTML key-rates page."""
    api_url = (
        "https://data-api.ecb.europa.eu/service/data/FM"
        "/B.U2.EUR.4F.KR.DFR.LEV?lastNObservations=1&format=jsondata"
    )
    log.info("Fetching EU rate (SDW API): %s", api_url)
    try:
        resp = requests.get(api_url, headers=HTTP_HEADERS, timeout=20)
        resp.raise_for_status()
        data = resp.json()
        series = data["dataSets"][0]["series"]
        obs = next(iter(series.values()))["observations"]
        rate = float(next(iter(obs.values()))[0])
        log.info("  -> EU rate (API): %.4f%%", rate)
        return rate
    except Exception as e:
        log.warning("  ECB API failed (%s), trying HTML", e)

    html_url = (
        "https://www.ecb.europa.eu/stats/policy_and_exchange_rates"
        "/key_ecb_interest_rates/html/index.en.html"
    )
    soup = fetch(html_url)
    tbody = soup.find("tbody")
    if tbody:
        row = tbody.find("tr")
        if row:
            for td in row.find_all("td"):
                s = td.find("strong")
                if s:
                    text = s.get_text(strip=True)
                    if re.match(r"^\d+\.\d+$", text):
                        rate = float(text)
                        if 0.0 <= rate <= 20.0:
                            log.info("  -> EU rate (HTML): %.4f%%", rate)
                            return rate
    raise ValueError("EU rate not found")


def scrape_boc_rate() -> float:
    """
    Bank of Canada #target-table. The most-recent row (first in tbody)
    whose date is not in the future gives the current overnight target.
    """
    url = "https://www.bankofcanada.ca/core-functions/monetary-policy/key-interest-rate/"
    log.info("Fetching CA rate: %s", url)
    soup  = fetch(url)
    today = date.today()
    table = soup.find("table", id="target-table")
    if table:
        for row in table.find_all("tr"):
            th = row.find("th", attrs={"data-date": True})
            if not th: continue
            try:
                dt = datetime.strptime(th["data-date"], "%Y-%m-%d").date()
            except ValueError: continue
            if dt > today: continue   # skip future scheduled dates
            tds = row.find_all("td")
            if tds:
                m = re.search(r"([\d.]+)", tds[0].get_text(strip=True))
                if m:
                    rate = float(m.group(1))
                    log.info("  -> CA rate: %.4f%%", rate)
                    return rate
    raise ValueError("CA rate not found")


def scrape_rba_rate() -> float:
    """
    RBA cash rate statistics page.
    The page has a table whose header contains 'Cash Rate Target'.
    We find that table, identify the target-rate column, and take the
    first data row (most recent decision).
    """
    url = "https://www.rba.gov.au/statistics/cash-rate/"
    log.info("Fetching AU rate: %s", url)
    soup = fetch(url)

    for table in soup.find_all("table"):
        full_text = table.get_text(" ", strip=True).lower()
        if "cash rate target" not in full_text:
            continue

        rows = table.find_all("tr")
        if len(rows) < 2:
            continue

        # Identify column index for "target" from the header row
        header_cells = rows[0].find_all(["th", "td"])
        target_col = None
        for i, cell in enumerate(header_cells):
            if "target" in cell.get_text(strip=True).lower():
                target_col = i
                break
        if target_col is None and len(header_cells) >= 2:
            target_col = 1  # fallback: second column

        # Walk data rows (most recent first) until we find a plausible rate
        for row in rows[1:10]:
            cells = row.find_all(["td", "th"])
            if target_col is not None and len(cells) > target_col:
                m = re.search(r"([\d.]+)", cells[target_col].get_text(strip=True))
            elif len(cells) >= 2:
                m = re.search(r"([\d.]+)", cells[1].get_text(strip=True))
            else:
                continue
            if m:
                rate = float(m.group(1))
                if 1.0 <= rate <= 15.0:   # plausible RBA cash-rate range
                    log.info("  -> AU rate: %.4f%%", rate)
                    return rate

    raise ValueError("AU rate not found")



# ─────────────────────────────────────────────────────────────────────────────
#  OIS / FUTURES — IMPLIED RATES
# ─────────────────────────────────────────────────────────────────────────────
# These functions return a list of implied policy rates (%) for each upcoming
# meeting.  On failure they return None and the caller falls back to the DRIFT
# model so the dashboard always has something to display.

# CME 30-day Fed Funds futures month codes (CBOT)
_FF_MONTH = {1:'F',2:'G',3:'H',4:'J',5:'K',6:'M',
             7:'N',8:'Q',9:'U',10:'V',11:'X',12:'Z'}


def _tradingview_ff_prices(tickers: list) -> "dict | None":
    """
    Fetch ZQ 30-Day Fed Funds Futures close prices via TradingView scanner API.
    tickers: list of TradingView symbols, e.g. ["CBOT:ZQK2026", "CBOT:ZQM2026"]
    Returns dict {symbol: price} or None on failure.
    """
    url = "https://scanner.tradingview.com/futures/scan"
    payload = {"symbols": {"tickers": tickers}, "columns": ["close"]}
    try:
        r = requests.post(url, json=payload, headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Content-Type": "application/json",
        }, timeout=15)
        r.raise_for_status()
        data = r.json()
        if data.get("error"):
            log.warning("TradingView error: %s", data["error"])
            return None
        result = {}
        for item in data.get("data", []):
            sym = item["s"]
            vals = item.get("d", [])
            if vals and vals[0] is not None:
                result[sym] = float(vals[0])
        return result if result else None
    except Exception as exc:
        log.debug("TradingView scanner: %s", exc)
    return None


def _yahoo_price(ticker: str) -> "float | None":
    """Return the most recent closing price from Yahoo Finance chart API (fallback)."""
    url = (f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}"
           f"?interval=1d&range=5d")
    try:
        resp = requests.get(url, headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept": "application/json",
        }, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        closes = (data["chart"]["result"][0]["indicators"]
                      ["quote"][0].get("close", []))
        for v in reversed(closes):
            if v is not None:
                return float(v)
    except Exception as exc:
        log.debug("Yahoo %s: %s", ticker, exc)
    return None


def fetch_us_implied_rates(meetings: list[str]) -> "list[float] | None":
    """
    Derive per-meeting implied Fed Funds rates from CME 30-day FF futures.

    Primary source: TradingView scanner API (CBOT:ZQ{M}{YY}, field "close").
      implied_rate = 100 − price.  No API key required.

    Fallback: Yahoo Finance ZQ{month_code}{YY}.CBT.

    For an FOMC meeting in month M the month M+1 contract captures the
    post-decision rate — mirrors the CME FedWatch methodology.
    """
    # Build per-meeting contract specs
    contract_specs = []
    for mtg_str in meetings:
        try:
            mtg = datetime.strptime(mtg_str, "%d %b %Y").date()
        except ValueError:
            log.warning("  US implied: bad meeting date: %s", mtg_str)
            return None
        post_month = mtg.month + 1
        post_year  = mtg.year
        if post_month > 12:
            post_month = 1
            post_year += 1
        tv_sym   = f"CBOT:ZQ{_FF_MONTH[post_month]}{post_year}"
        yf_tick  = f"ZQ{_FF_MONTH[post_month]}{str(post_year)[-2:]}.CBT"
        contract_specs.append((mtg_str, tv_sym, yf_tick))

    # ── Primary: TradingView batch fetch ─────────────────────────────────────
    tv_tickers = [s[1] for s in contract_specs]
    log.info("  US implied: TradingView batch — %s", tv_tickers)
    tv_prices  = _tradingview_ff_prices(tv_tickers)

    results: list[float] = []
    for mtg_str, tv_sym, yf_tick in contract_specs:
        price: "float | None" = None

        if tv_prices and tv_sym in tv_prices:
            price = tv_prices[tv_sym]
            implied = round(100.0 - price, 4)
            log.info("  US implied[%s]: TradingView %s=%.4f → %.4f%%",
                     mtg_str, tv_sym, price, implied)
            results.append(implied)
            continue

        # ── Fallback: Yahoo Finance ───────────────────────────────────────────
        log.warning("  US implied[%s]: TradingView %s missing — trying Yahoo %s",
                    mtg_str, tv_sym, yf_tick)
        price = _yahoo_price(yf_tick)
        if price is None:
            log.warning("  US implied: no data for %s — aborting", yf_tick)
            return None
        implied = round(100.0 - price, 4)
        log.info("  US implied[%s]: Yahoo %s=%.4f → %.4f%%",
                 mtg_str, yf_tick, price, implied)
        results.append(implied)

    if len(results) != len(meetings):
        return None
    return [(r, False) for r in results]  # ZQ: every meeting has a direct contract


def _interpolate_curve(
    sorted_curve: list[tuple[float, float]],
    meetings:     list[str],
    label:        str,
) -> "list[float] | None":
    """
    Linearly interpolate per-meeting implied rates from a term-structure curve.

    sorted_curve : [(years_from_today, rate_pct), ...] sorted ascending by tenor
    meetings     : meeting date strings in "%d %b %Y" format
    label        : log prefix (e.g. "ECB", "UK OIS")
    """
    today = date.today()
    results: list[float] = []
    for mtg_str in meetings:
        try:
            mtg = datetime.strptime(mtg_str, "%d %b %Y").date()
        except ValueError:
            return None
        t = max((mtg - today).days / 365.0, 0.01)

        if t <= sorted_curve[0][0]:
            rate = sorted_curve[0][1]
        elif t >= sorted_curve[-1][0]:
            rate = sorted_curve[-1][1]
        else:
            rate = sorted_curve[-1][1]
            for j in range(len(sorted_curve) - 1):
                t1, r1 = sorted_curve[j]
                t2, r2 = sorted_curve[j + 1]
                if t1 <= t <= t2:
                    rate = r1 + (r2 - r1) * (t - t1) / (t2 - t1)
                    break

        results.append(round(rate, 4))
        log.info("  %s implied[%s]: t=%.3fy → %.4f%%", label, mtg_str, t, rate)

    return results if len(results) == len(meetings) else None


# ── EU: ECB — 3-Month EURIBOR futures (Eurex FEU3) ────────────────────────────

# Eurex FEU3 month codes (same convention as CME ZQ)
_EURIBOR_MONTH = {1:'F',2:'G',3:'H',4:'J',5:'K',6:'M',
                  7:'N',8:'Q',9:'U',10:'V',11:'X',12:'Z'}

def fetch_ecb_implied_rates(meetings: list[str]) -> "list[float] | None":
    """
    ECB implied rates derived from Eurex 3-Month EURIBOR futures (FEU3).

    implied_euribor = 100 − price
    implied_ecb     = implied_euribor − spread

    The EURIBOR-OIS spread is calibrated dynamically: the nearest-dated contract
    minus the current ECB DFR.  This anchors the curve at DFR for near-term
    meetings and lets subsequent contracts reflect the market-implied path.

    For each meeting in month M the contract expiring in month M is used.
    If that contract is unavailable the nearest available contract is used
    via linear interpolation.

    Source: TradingView scanner API, exchange EUREX.
    Fallback: ECB SDW AAA government bond yield curve (SR_3M … SR_3Y).
    """
    today = date.today()

    # ── Build list of contracts needed ───────────────────────────────────────
    contract_months: list[tuple[str, str, float]] = []  # (mtg_str, tv_sym, t_yr)
    for mtg_str in meetings:
        try:
            mtg = datetime.strptime(mtg_str, "%d %b %Y").date()
        except ValueError:
            return None
        t = max((mtg - today).days / 365.0, 0.001)
        tv_sym = f"EUREX:FEU3{_EURIBOR_MONTH[mtg.month]}{mtg.year}"
        contract_months.append((mtg_str, tv_sym, t))

    # Also fetch the nearest contract (for anchoring), which may differ from
    # the first meeting's month if the first meeting has already passed or is
    # in the same month as an already-expired contract.
    anchor_sym = f"EUREX:FEU3{_EURIBOR_MONTH[today.month]}{today.year}"

    all_syms = list(dict.fromkeys(
        [anchor_sym] + [s for _, s, _ in contract_months]
    ))

    log.info("  EU implied: EURIBOR futures batch — %s", all_syms)
    tv_prices = _tradingview_ff_prices(all_syms)  # reuse the same batch fetcher

    if not tv_prices or len(tv_prices) < 2:
        log.warning("  EU implied: TradingView returned <2 contracts — falling back to ECB YC")
        fb = _ecb_yc_fallback(meetings)
        return fb  # already returns tuples

    # ── Compute implied EURIBOR rates for each contract we got ───────────────
    euribor_curve: list[tuple[float, float]] = []  # (t_yr, implied_euribor)
    for sym, price in tv_prices.items():
        # Parse year and month from symbol, e.g. EUREX:FEU3M2026 → M, 2026
        code = sym[len("EUREX:FEU3"):]   # e.g. "M2026"
        month_code = code[0]
        year = int(code[1:])
        month = next(m for m, c in _EURIBOR_MONTH.items() if c == month_code)
        t = max((date(year, month, 15) - today).days / 365.0, 0.001)
        euribor_curve.append((t, round(100.0 - price, 4)))

    euribor_curve.sort()

    # ── Anchor: spread = nearest contract's implied EURIBOR − current DFR ───
    ecb_rate = scrape_ecb_rate()
    nearest_euribor = euribor_curve[0][1]
    spread = nearest_euribor - ecb_rate
    log.info("  EU implied: EURIBOR-OIS spread (calibrated) = %.2fbps",
             spread * 100)

    # ── Build DFR-anchored curve ──────────────────────────────────────────────
    adjusted = [(t, r - spread) for t, r in euribor_curve]

    rates = _interpolate_curve(adjusted, meetings, "ECB EURIBOR")
    if rates is None:
        return None
    # Mark meeting as direct if its month had a contract in tv_prices
    available = set()
    for sym in tv_prices:
        mc = sym[len("EUREX:FEU3")]  # single character
        yr = int(sym[len("EUREX:FEU3") + 1:])
        available.add((yr, next(m for m, c in _EURIBOR_MONTH.items() if c == mc)))
    return [(r, (datetime.strptime(mtg, "%d %b %Y").month,
                 datetime.strptime(mtg, "%d %b %Y").year) not in
             {(m, y) for y, m in available})
            for r, mtg in zip(rates, meetings)]


def _ecb_yc_fallback(meetings: list[str]) -> "list[float] | None":
    """
    Fallback: ECB SDW AAA government bond yield curve (SR_3M … SR_3Y).
    Flat below 3M — used only when EURIBOR futures are unavailable.
    """
    today = date.today()
    tenor_map = {"3M": 0.25, "6M": 0.5, "9M": 0.75, "2Y": 2.0, "3Y": 3.0}
    raw: dict[float, float] = {}

    for tenor_code, tenor_yr in tenor_map.items():
        url = (f"https://data-api.ecb.europa.eu/service/data/YC"
               f"/B.U2.EUR.4F.G_N_A.SV_C_YM.SR_{tenor_code}"
               f"?lastNObservations=1&format=jsondata")
        try:
            resp = requests.get(
                url, headers={"Accept": "application/json", **HTTP_HEADERS},
                timeout=12,
            )
            if resp.status_code != 200:
                continue
            data   = resp.json()
            series = data["dataSets"][0]["series"]
            obs    = next(iter(series.values()))["observations"]
            values = [(int(k), v[0]) for k, v in obs.items() if v[0] is not None]
            if values:
                raw[tenor_yr] = float(sorted(values)[-1][1])
                log.info("  ECB YC SR_%s: %.4f%%", tenor_code, raw[tenor_yr])
        except Exception as exc:
            log.debug("  ECB YC SR_%s: %s", tenor_code, exc)

    if len(raw) < 2:
        return None

    ecb_rate = scrape_ecb_rate()
    base_3m  = raw.get(0.25, min(raw.values()))
    adjusted = sorted([(t, ecb_rate + (r - base_3m)) for t, r in raw.items()])
    rates = _interpolate_curve(adjusted, meetings, "ECB YC")
    if rates is None:
        return None
    return [(r, True) for r in rates]  # YC fallback: all interpolated from smooth curve


# ── UK: Bank of England — 3-Month SONIA futures (ICE) ────────────────────────

def fetch_uk_implied_rates(meetings: list[str]) -> "list | None":
    """
    Primary: ICE 3-Month SONIA futures (ICEEUR:SO3{M}{YYYY}) via TradingView.
    Same anchoring approach as ECB EURIBOR: spread = nearest_contract - BOE rate.
    Fallback: BOE SONIA OIS instantaneous forward curve (ZIP, daily).
    """
    today   = date.today()
    boe_rate = scrape_boe_rate()

    anchor_sym  = f"ICEEUR:SO3{_FF_MONTH[today.month]}{today.year}"
    meeting_syms = [(mtg, f"ICEEUR:SO3{_FF_MONTH[datetime.strptime(mtg,'%d %b %Y').month]}{datetime.strptime(mtg,'%d %b %Y').year}")
                    for mtg in meetings]

    all_syms = list(dict.fromkeys([anchor_sym] + [s for _, s in meeting_syms]))
    log.info("  UK implied: SONIA futures batch — %s", all_syms)
    tv_prices = _tradingview_ff_prices(all_syms)

    if tv_prices and len(tv_prices) >= 2:
        sonia_curve = []
        for sym, price in tv_prices.items():
            code = sym[len("ICEEUR:SO3"):]
            mc, yr = code[0], int(code[1:])
            month = next(m for m, c in _FF_MONTH.items() if c == mc)
            t = max((date(yr, month, 15) - today).days / 365.0, 0.001)
            sonia_curve.append((t, round(100.0 - price, 4)))
        sonia_curve.sort()

        spread = sonia_curve[0][1] - boe_rate
        log.info("  UK implied: SONIA-Bank Rate spread (calibrated) = %.2fbps", spread * 100)
        adjusted = [(t, r - spread) for t, r in sonia_curve]

        available = set()
        for sym in tv_prices:
            code = sym[len("ICEEUR:SO3"):]
            mc, yr = code[0], int(code[1:])
            available.add((yr, next(m for m, c in _FF_MONTH.items() if c == mc)))

        rates = _interpolate_curve(adjusted, meetings, "UK SONIA")
        if rates is not None:
            return [(r, (datetime.strptime(mtg, "%d %b %Y").year,
                         datetime.strptime(mtg, "%d %b %Y").month) not in available)
                    for r, mtg in zip(rates, meetings)]

    log.warning("  UK implied: TradingView SONIA failed — falling back to BOE OIS")
    fb = _boe_ois_fallback(meetings)
    return [(r, True) for r in fb] if fb else None


def _boe_ois_fallback(meetings: list[str]) -> "list[float] | None":
    """BOE SONIA OIS instantaneous forward curve fallback (returns plain floats)."""
    if not _OPENPYXL:
        log.warning("  UK implied: openpyxl not installed — using DRIFT")
        return None

    import io, zipfile
    today = date.today()

    # Try current-month file first; fall back to annual history file
    for zip_url in [
        "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/latest-yield-curve-data.zip",
        "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/oisddata.zip",
    ]:
        try:
            resp = requests.get(zip_url, headers=HTTP_HEADERS, timeout=30)
            resp.raise_for_status()
            z = zipfile.ZipFile(io.BytesIO(resp.content))
            # Find the OIS file in the archive
            ois_name = next(
                (n for n in z.namelist() if "OIS" in n and n.endswith(".xlsx")),
                None,
            )
            if ois_name is None:
                continue
            with z.open(ois_name) as f:
                wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
            ws = wb["1. fwds, short end"]

            # Row 3 = maturity months (approx integers); row 4 = maturity years
            n_cols = ws.max_column
            tenor_yrs = {}   # col_index → years
            for c in range(2, n_cols + 1):
                yr_val = ws.cell(4, c).value
                if isinstance(yr_val, (int, float)) and yr_val > 0:
                    tenor_yrs[c] = float(yr_val)

            # Find most recent data row (col A = datetime)
            latest_row: dict[float, float] = {}
            for row in range(ws.max_row, 5, -1):
                dt = ws.cell(row, 1).value
                if not isinstance(dt, datetime):
                    continue
                # Collect non-null values
                for c, yr in tenor_yrs.items():
                    val = ws.cell(row, c).value
                    if isinstance(val, (int, float)):
                        latest_row[yr] = float(val)
                if len(latest_row) >= 3:
                    log.info("  UK OIS: using date %s (%d tenors)",
                             dt.date(), len(latest_row))
                    break

            if len(latest_row) < 3:
                log.warning("  UK OIS: %s had < 3 valid rows", ois_name)
                continue

            sorted_curve = sorted(latest_row.items())
            return _interpolate_curve(sorted_curve, meetings, "UK OIS")

        except Exception as exc:
            log.warning("  UK OIS (%s): %s", zip_url.split("/")[-1], exc)

    log.warning("  UK implied: all sources failed — using DRIFT")
    return None


# ── CA: Bank of Canada — 3-Month CORRA futures (TMX, quarterly) ──────────────

# CORRA futures only have quarterly contracts (H=Mar, M=Jun, U=Sep, Z=Dec)
_CORRA_QUARTERLY = {3, 6, 9, 12}

def fetch_ca_implied_rates(meetings: list[str]) -> "list | None":
    """
    Primary: TMX 3-Month CORRA futures (TMX:CRA{M}{YYYY}) via TradingView.
    Quarterly contracts (Mar/Jun/Sep/Dec) only — non-quarterly meetings are
    interpolated and flagged with is_interp=True.
    Same anchoring approach as ECB EURIBOR.
    Fallback: BOC Valet CORRA overnight + 2Y bond interpolation.
    """
    today    = date.today()
    boc_rate = scrape_boc_rate()

    # Request the nearest 6 quarterly contracts spanning the meeting window
    quarterly_codes = {3: 'H', 6: 'M', 9: 'U', 12: 'Z'}
    syms = []
    for yr in [today.year, today.year + 1]:
        for m, mc in quarterly_codes.items():
            sym = f"TMX:CRA{mc}{yr}"
            if date(yr, m, 15) >= today - __import__('datetime').timedelta(days=30):
                syms.append(sym)
    syms = syms[:8]  # up to 2 years out

    log.info("  CA implied: CORRA futures batch — %s", syms)
    tv_prices = _tradingview_ff_prices(syms)

    if tv_prices and len(tv_prices) >= 2:
        corra_curve = []
        for sym, price in tv_prices.items():
            code = sym[len("TMX:CRA"):]
            mc, yr = code[0], int(code[1:])
            month = next(m for m, c in quarterly_codes.items() if c == mc)
            t = max((date(yr, month, 15) - today).days / 365.0, 0.001)
            corra_curve.append((t, round(100.0 - price, 4)))
        corra_curve.sort()

        spread = corra_curve[0][1] - boc_rate
        log.info("  CA implied: CORRA-BOC spread (calibrated) = %.2fbps", spread * 100)
        adjusted = [(t, r - spread) for t, r in corra_curve]

        available_qtr = set()
        for sym in tv_prices:
            code = sym[len("TMX:CRA"):]
            mc, yr = code[0], int(code[1:])
            available_qtr.add((yr, next(m for m, c in quarterly_codes.items() if c == mc)))

        rates = _interpolate_curve(adjusted, meetings, "CA CORRA")
        if rates is not None:
            result = []
            for r, mtg in zip(rates, meetings):
                mtg_d = datetime.strptime(mtg, "%d %b %Y")
                is_interp = (mtg_d.year, mtg_d.month) not in available_qtr
                result.append((r, is_interp))
            return result

    log.warning("  CA implied: TradingView CORRA failed — falling back to BOC CORRA+2Y")
    fb = _boc_corra_2y_fallback(meetings)
    return [(r, True) for r in fb] if fb else None


def _boc_corra_2y_fallback(meetings: list[str]) -> "list[float] | None":
    """BOC Valet CORRA overnight + 2Y bond fallback (returns plain floats)."""
    today = date.today()
    valet = "https://www.bankofcanada.ca/valet/observations/{}/json?recent=5"

    def boc_latest(series: str) -> "float | None":
        try:
            resp = requests.get(
                valet.format(series), headers=HTTP_HEADERS, timeout=12
            )
            resp.raise_for_status()
            obs = resp.json().get("observations", [])
            for row in reversed(obs):
                v = row.get(series, {})
                val = v.get("v") if isinstance(v, dict) else v
                if val not in (None, ""):
                    return float(val)
        except Exception as exc:
            log.debug("  CA BOC %s: %s", series, exc)
        return None

    corra = boc_latest("V122514")          # overnight CORRA
    bond2y = boc_latest("BD.CDN.2YR.DQ.YLD")  # 2Y GoC bond

    if corra is None or bond2y is None:
        log.warning("  CA implied: BOC API incomplete (CORRA=%s, 2Y=%s) — "
                    "using DRIFT", corra, bond2y)
        return None

    log.info("  CA BOC: CORRA=%.4f%%, 2Y=%.4f%%", corra, bond2y)

    # Build simple 2-point curve: [overnight anchor, 2-year]
    # The overnight anchor is placed at a very short horizon (3 business days)
    sorted_curve = sorted([(0.01, corra), (2.0, bond2y)])
    return _interpolate_curve(sorted_curve, meetings, "CA BOC")


# ── AU: Reserve Bank of Australia — ASX 30-day IB cash rate futures ──────────

_BBSW_OIS_SPREAD = {1/12: 0.12, 3/12: 0.22, 6/12: 0.30}  # kept for BABs fallback

def fetch_au_implied_rates(meetings: list[str]) -> "list | None":
    """
    Primary: ASX 30-day Interbank Cash Rate futures (ASX24:IB{M}{YYYY}).
    Settlement = average RBA cash rate over the calendar month.
    Convention: use month M+1 contract for a meeting in month M (same as ZQ).
    No spread adjustment needed — these settle directly to the RBA cash rate.
    Fallback: RBA F01 BABs/NCDs with BBSW-OIS spread adjustment.
    """
    today    = date.today()

    contract_specs = []
    for mtg_str in meetings:
        try:
            mtg = datetime.strptime(mtg_str, "%d %b %Y").date()
        except ValueError:
            return None
        post_month = mtg.month + 1
        post_year  = mtg.year
        if post_month > 12:
            post_month = 1
            post_year += 1
        tv_sym = f"ASX24:IB{_FF_MONTH[post_month]}{post_year}"
        contract_specs.append((mtg_str, tv_sym))

    all_syms = list(dict.fromkeys(s for _, s in contract_specs))
    log.info("  AU implied: IB cash rate futures batch — %s", all_syms)
    tv_prices = _tradingview_ff_prices(all_syms)

    if tv_prices and len(tv_prices) >= 2:
        results = []
        for mtg_str, tv_sym in contract_specs:
            if tv_sym in tv_prices:
                implied = round(100.0 - tv_prices[tv_sym], 4)
                log.info("  AU IB implied[%s]: %s=%.4f%% -> %.4f%%",
                         mtg_str, tv_sym, tv_prices[tv_sym], implied)
                results.append((implied, False))
            else:
                # Symbol missing — interpolate from adjacent contracts
                log.warning("  AU implied[%s]: %s missing — interpolating", mtg_str, tv_sym)
                # Build curve from what we have and interpolate
                curve = []
                for sym, price in tv_prices.items():
                    code = sym[len("ASX24:IB"):]
                    mc, yr = code[0], int(code[1:])
                    month = next(m for m, c in _FF_MONTH.items() if c == mc)
                    t = max((date(yr, month, 15) - today).days / 365.0, 0.001)
                    curve.append((t, round(100.0 - price, 4)))
                curve.sort()
                mtg_d = datetime.strptime(mtg_str, "%d %b %Y").date()
                t_mtg = max((mtg_d - today).days / 365.0, 0.001)
                interped = _interpolate_curve(curve, [mtg_str], "AU IB interp")
                results.append((interped[0] if interped else None, True))
        if all(r is not None for r, _ in results):
            return results

    log.warning("  AU implied: TradingView IB failed — falling back to RBA BABs")
    fb = _rba_babs_fallback(meetings)
    return [(r, True) for r in fb] if fb else None


def _rba_babs_fallback(meetings: list[str]) -> "list[float] | None":
    """RBA F01 BABs/NCDs BBSW-OIS-adjusted fallback (returns plain floats)."""
    if not _OPENPYXL:
        log.warning("  AU implied: openpyxl not installed — using DRIFT")
        return None

    import io
    today = date.today()
    url   = "https://www.rba.gov.au/statistics/tables/xls/f01d.xlsx"
    try:
        resp = requests.get(url, headers=HTTP_HEADERS, timeout=20)
        resp.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
        ws = wb.active
        # Columns: 10=BAB 1M, 11=BAB 3M, 12=BAB 6M (0-indexed: rows start at 12)
        # Find most recent row with non-empty BABs
        babs: dict[float, float] = {}
        for row in range(ws.max_row, 10, -1):
            b1 = ws.cell(row, 10).value
            b3 = ws.cell(row, 11).value
            b6 = ws.cell(row, 12).value
            dt = ws.cell(row, 1).value
            if b3 not in (None, "") and isinstance(b3, (int, float)):
                log.info("  AU RBA BABs: date=%s, 1M=%.4f%%, 3M=%.4f%%, 6M=%.4f%%",
                         dt, b1, b3, b6 if b6 else 0)
                if b1 not in (None, ""):
                    babs[1/12] = float(b1) - _BBSW_OIS_SPREAD[1/12]
                babs[3/12]     = float(b3) - _BBSW_OIS_SPREAD[3/12]
                if b6 not in (None, ""):
                    babs[6/12] = float(b6) - _BBSW_OIS_SPREAD[6/12]
                break

        if len(babs) < 2:
            log.warning("  AU implied: BABs data missing — using DRIFT")
            return None

        sorted_curve = sorted(babs.items())
        return _interpolate_curve(sorted_curve, meetings, "AU BABs")

    except Exception as exc:
        log.warning("  AU implied: %s — using DRIFT", exc)
        return None


# Registry of implied-rate fetchers per market
# (keyed by market code; returns list[float] or None)
IMPLIED_RATE_FETCHERS: dict = {
    "US": fetch_us_implied_rates,
    "EU": fetch_ecb_implied_rates,
    "UK": fetch_uk_implied_rates,
    "CA": fetch_ca_implied_rates,
    "AU": fetch_au_implied_rates,
}


# ─────────────────────────────────────────────────────────────────────────────
#  ORCHESTRATION
# ─────────────────────────────────────────────────────────────────────────────

LIVE_SCRAPERS = {
    "meetings": {
        "US": scrape_fed_meetings,
        "EU": scrape_ecb_meetings,
        "UK": scrape_boe_meetings,
        "CA": scrape_boc_meetings,
        "AU": scrape_rba_meetings,
    },
    "rates": {
        "US": scrape_fed_rate,
        "EU": scrape_ecb_rate,
        "UK": scrape_boe_rate,
        "CA": scrape_boc_rate,
        "AU": scrape_rba_rate,
    },
}

ALL_MARKETS = list(META.keys())


def load_saved_data() -> dict:
    if os.path.exists(JSON_FILE):
        try:
            with open(JSON_FILE, "r", encoding="utf-8") as f:
                return json.load(f).get("markets", {})
        except Exception:
            pass
    return {k: dict(v) for k, v in FALLBACK.items()}


def fetch_all_data() -> tuple[dict, list[str]]:
    saved  = load_saved_data()
    result = {code: dict(saved.get(code, {})) for code in ALL_MARKETS}
    errors: list[str] = []

    for code in ALL_MARKETS:
        # ── meetings ────────────────────────────────────────────────────────
        if code in LIVE_SCRAPERS["meetings"]:
            try:
                result[code]["meetings"] = LIVE_SCRAPERS["meetings"][code]()
            except Exception as exc:
                msg = f"{code} meetings: {exc}"
                errors.append(msg)
                log.warning("  !! %s  (using fallback)", msg)
                if "meetings" not in result[code]:
                    result[code]["meetings"] = FALLBACK[code]["meetings"]
        else:
            # Hardcoded markets — always use fallback (most recent data from FALLBACK)
            if "meetings" not in result[code] or not result[code]["meetings"]:
                result[code]["meetings"] = FALLBACK[code]["meetings"]
            log.info("  -> %s meetings: using hardcoded fallback (%d dates)",
                     code, len(result[code]["meetings"]))

        # ── rates ────────────────────────────────────────────────────────────
        if code in LIVE_SCRAPERS["rates"]:
            try:
                result[code]["rate"] = LIVE_SCRAPERS["rates"][code]()
            except Exception as exc:
                msg = f"{code} rate: {exc}"
                errors.append(msg)
                log.warning("  !! %s  (using fallback)", msg)
                if "rate" not in result[code]:
                    result[code]["rate"] = FALLBACK[code]["rate"]
        else:
            if "rate" not in result[code]:
                result[code]["rate"] = FALLBACK[code]["rate"]
            log.info("  -> %s rate: using hardcoded fallback (%.4f%%)",
                     code, result[code]["rate"])

        # ── static metadata + drift ───────────────────────────────────────────
        result[code].update(META[code])

        # ── implied rates (OIS / futures) ─────────────────────────────────────
        if code in IMPLIED_RATE_FETCHERS:
            try:
                implied = IMPLIED_RATE_FETCHERS[code](result[code]["meetings"])
                if implied:
                    rates  = [r for r, _ in implied]
                    interp = [i for _, i in implied]
                    result[code]["impliedRates"]       = rates
                    result[code]["impliedRatesInterp"] = interp
                    log.info("  %s impliedRates: %s", code,
                             [f"{r:.3f}" for r in rates])
                else:
                    log.info("  %s impliedRates: fetcher returned None — "
                             "dashboard will use DRIFT model", code)
            except Exception as exc:
                log.warning("  %s impliedRates failed (%s) — using DRIFT", code, exc)

    return result, errors


# ─────────────────────────────────────────────────────────────────────────────
#  HISTORY
# ─────────────────────────────────────────────────────────────────────────────

def _hist_sort_key(date_str: str) -> str:
    """Convert 'DD Mon YYYY' to 'YYYY-MM-DD' for sorting."""
    try:
        return datetime.strptime(date_str, "%d %b %Y").strftime("%Y-%m-%d")
    except Exception:
        return date_str


def load_history() -> dict:
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as exc:
            log.warning("History load failed: %s", exc)
    return {}


def save_history(history: dict) -> None:
    try:
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, separators=(",", ":"))
        log.info("History saved -> %s", HISTORY_FILE)
    except Exception as exc:
        log.warning("History save failed: %s", exc)


def _trim_sort(snapshots: list) -> list:
    """Remove entries older than HISTORY_MAX_DAYS and sort ascending."""
    from datetime import timedelta
    cutoff = (date.today() - timedelta(days=HISTORY_MAX_DAYS)).strftime("%Y-%m-%d")
    kept = [s for s in snapshots if _hist_sort_key(s.get("date", "")) >= cutoff]
    kept.sort(key=lambda s: _hist_sort_key(s.get("date", "")))
    return kept


def fetch_us_history_yfinance(meetings: list) -> list:
    """
    Backfill US pricing history via yfinance ZQ futures.
    Returns [{date: 'DD Mon YYYY', impliedRates: [r0, r1, ...]}, ...] sorted oldest→newest.
    """
    try:
        import yfinance as yf
    except ImportError:
        log.warning("  US hist: yfinance not installed — run: pip install yfinance")
        return []

    from datetime import timedelta
    start = (date.today() - timedelta(days=HISTORY_MAX_DAYS)).strftime("%Y-%m-%d")
    by_date: dict = {}

    for i, mtg_str in enumerate(meetings):
        try:
            mtg = datetime.strptime(mtg_str, "%d %b %Y")
            post_month = mtg.month + 1
            post_year = mtg.year
            if post_month > 12:
                post_month = 1
                post_year += 1
            mc = _FF_MONTH[post_month]
            ticker = f"ZQ{mc}{str(post_year)[2:]}.CBT"
            df = yf.download(ticker, start=start, progress=False, auto_adjust=True)
            # yfinance may return MultiIndex columns; flatten
            if getattr(df.columns, 'nlevels', 1) > 1:
                df.columns = df.columns.get_level_values(0)
            if df.empty:
                log.warning("  US hist: no data for %s", ticker)
                continue
            for idx, row in df.iterrows():
                d = idx.date() if hasattr(idx, "date") else idx
                ds = d.strftime("%d %b %Y")
                if ds not in by_date:
                    by_date[ds] = [None] * len(meetings)
                try:
                    by_date[ds][i] = round(100.0 - float(row["Close"]), 4)
                except Exception:
                    pass
            log.info("  US hist: %d pts for %s (%s)", len(df), ticker, mtg_str)
        except Exception as exc:
            log.warning("  US hist: failed for meeting %s: %s", mtg_str, exc)

    snaps = [
        {"date": ds, "impliedRates": r}
        for ds, r in sorted(by_date.items())
        if any(v is not None for v in r)
    ]
    log.info("  US hist yfinance: %d daily snapshots assembled", len(snaps))
    return snaps


def downsample_for_html(snapshots: list) -> list:
    """
    Reduce snapshot density for HTML injection to keep page size reasonable.
    Keeps: all of last 90 days daily, then every 7th (weekly) before that.
    """
    from datetime import timedelta
    cutoff_daily = (date.today() - timedelta(days=90)).strftime("%Y-%m-%d")
    result = []
    sparse_counter = 0
    for s in snapshots:  # already sorted oldest→newest
        key = _hist_sort_key(s.get("date", ""))
        if key >= cutoff_daily:
            result.append(s)
        else:
            if sparse_counter % 7 == 0:
                result.append(s)
            sparse_counter += 1
    return result


def fetch_eu_history_ecb(meetings: list) -> list:
    """
    Backfill EU ECB history using ECB SDW yield curve spot rates.
    Fetches daily 3M/6M/1Y/2Y/3Y spot rates and ECB DFR, interpolates per meeting.
    Returns [{date: 'DD Mon YYYY', impliedRates: [r0, r1, ...]}, ...].
    """
    from datetime import timedelta
    start = (date.today() - timedelta(days=HISTORY_MAX_DAYS)).strftime("%Y-%m-%d")

    # ── Fetch ECB DFR history ─────────────────────────────────────────────────
    dfr_by_date: dict = {}
    try:
        url = (f"https://data-api.ecb.europa.eu/service/data/FM/B.U2.EUR.4F.KR.DFR.LEV"
               f"?startPeriod={start}&format=jsondata")
        r = requests.get(url, timeout=30, headers=HTTP_HEADERS)
        r.raise_for_status()
        data = r.json()
        dim_vals = data["structure"]["dimensions"]["observation"][0]["values"]
        obs = list(data["dataSets"][0]["series"].values())[0]["observations"]
        for k, v in obs.items():
            if v[0] is not None:
                dfr_by_date[dim_vals[int(k)]["id"]] = float(v[0])
        log.info("  EU hist: DFR history — %d dates", len(dfr_by_date))
    except Exception as exc:
        log.warning("  EU hist: DFR history failed: %s", exc)
        return []

    # ── Fetch yield curve spot rates at 5 tenors ──────────────────────────────
    tenor_map = {"SR_3M": 0.25, "SR_6M": 0.5, "SR_1Y": 1.0, "SR_2Y": 2.0, "SR_3Y": 3.0}
    yc_by_date: dict = {}  # "YYYY-MM-DD" -> {tenor_years: rate}
    for tid, tyears in tenor_map.items():
        try:
            url = (f"https://data-api.ecb.europa.eu/service/data/YC/"
                   f"B.U2.EUR.4F.G_N_A.SV_C_YM.{tid}"
                   f"?startPeriod={start}&format=jsondata")
            r = requests.get(url, timeout=30, headers=HTTP_HEADERS)
            r.raise_for_status()
            data = r.json()
            dim_vals = data["structure"]["dimensions"]["observation"][0]["values"]
            obs = list(data["dataSets"][0]["series"].values())[0]["observations"]
            for k, v in obs.items():
                if v[0] is not None:
                    ds = dim_vals[int(k)]["id"]
                    yc_by_date.setdefault(ds, {})[tyears] = float(v[0])
        except Exception as exc:
            log.warning("  EU hist: YC %s failed: %s", tid, exc)

    if not yc_by_date:
        log.warning("  EU hist: no yield curve data retrieved")
        return []
    log.info("  EU hist: yield curve data — %d dates", len(yc_by_date))

    # ── Build per-date snapshots ───────────────────────────────────────────────
    ecb_meetings = [datetime.strptime(m, "%d %b %Y") for m in meetings]
    snaps = []
    last_dfr = None
    for ds in sorted(yc_by_date.keys()):
        yc = yc_by_date[ds]
        if len(yc) < 2:
            continue
        dfr = dfr_by_date.get(ds, last_dfr)
        if dfr is None:
            continue
        last_dfr = dfr
        snap_dt = datetime.strptime(ds, "%Y-%m-%d")
        sorted_curve = sorted(yc.items())
        base_3m = yc.get(0.25, sorted_curve[0][1])

        implied_rates = []
        for mtg in ecb_meetings:
            t = max((mtg - snap_dt).days / 365.0, 0.001)
            # Linear interpolate on the spot rate curve
            rate = sorted_curve[-1][1]  # default: use longest tenor
            if t <= sorted_curve[0][0]:
                rate = dfr  # very near-term: use policy rate
            else:
                for i in range(len(sorted_curve) - 1):
                    t0, r0 = sorted_curve[i]
                    t1, r1 = sorted_curve[i + 1]
                    if t0 <= t <= t1:
                        alpha = (t - t0) / max(t1 - t0, 0.001)
                        rate = r0 + alpha * (r1 - r0)
                        break
            # Anchor to DFR (same as ecb fallback)
            anchored = round(dfr + (rate - base_3m), 4)
            implied_rates.append(anchored)

        d_fmt = snap_dt.strftime("%d %b %Y")
        snaps.append({"date": d_fmt, "impliedRates": implied_rates})

    log.info("  EU hist ECB: %d snapshots assembled", len(snaps))
    return snaps


def fetch_uk_history_boe(meetings: list) -> list:
    """
    UK historical backfill placeholder.
    BoE IADB requires browser session; SONIA futures history is not freely available.
    History accumulates from hourly snapshots.
    """
    return []



# ─────────────────────────────────────────────────────────────────────────────
#  OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

def save_json(markets: dict, timestamp: str) -> None:
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump({"markets": markets, "last_updated": timestamp}, f, indent=2)
    log.info("Saved -> %s", JSON_FILE)


def inject_html(markets: dict, timestamp: str) -> bool:
    if not os.path.exists(HTML_FILE):
        log.error("wirp.html not found: %s", HTML_FILE)
        return False
    with open(HTML_FILE, "r", encoding="utf-8") as f:
        html = f.read()
    bi = html.find(DATA_BEGIN)
    ei = html.find(DATA_END)
    if bi == -1 or ei == -1:
        log.error("Data markers not found in wirp.html")
        return False
    compact = json.dumps(markets, separators=(",", ":"), ensure_ascii=False)
    block = (
        f"{DATA_BEGIN}\n"
        f"<script>\n"
        f"const MARKET_DATA={compact};\n"
        f'const LAST_UPDATED="{timestamp}";\n'
        f"</script>\n"
        f"{DATA_END}"
    )
    updated = html[:bi] + block + html[ei + len(DATA_END):]
    with open(HTML_FILE, "w", encoding="utf-8") as f:
        f.write(updated)
    log.info("Injected -> %s", HTML_FILE)
    return True


# ─────────────────────────────────────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="WIRP data updater — G10 + Asia")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print scraped data without writing files")
    args = parser.parse_args()

    log.info("=" * 60)
    log.info("WIRP updater starting  (dry-run=%s)", args.dry_run)

    markets, errors = fetch_all_data()
    timestamp = datetime.now(timezone.utc).strftime("%d %b %Y %H:%M UTC")

    if args.dry_run:
        print("\n-- Scraped data --------------------------------------------------")
        print(json.dumps({"markets": markets, "last_updated": timestamp}, indent=2))
        if errors:
            print(f"\n-- {len(errors)} error(s) --")
            for e in errors: print(f"  * {e}")
        return

    # ── History management ────────────────────────────────────────────────────
    history = load_history()
    today_str = date.today().strftime("%d %b %Y")

    for code in ALL_MARKETS:
        if code not in history:
            history[code] = []

        impl = markets.get(code, {}).get("impliedRates")
        if impl:
            # Upsert today's snapshot
            history[code] = [s for s in history[code] if s.get("date") != today_str]
            history[code].append({"date": today_str, "impliedRates": impl})

        # Backfill from external sources if history is sparse (< 30 entries = first run)
        if len(history[code]) < 30:
            backfill = []
            if code == "US":
                log.info("  US: sparse history — running yfinance backfill...")
                backfill = fetch_us_history_yfinance(markets["US"]["meetings"])
            elif code == "EU":
                log.info("  EU: sparse history — running ECB SDW backfill...")
                backfill = fetch_eu_history_ecb(markets["EU"]["meetings"])
            elif code == "UK":
                log.info("  UK: sparse history — running BoE backfill...")
                backfill = fetch_uk_history_boe(markets["UK"]["meetings"])
            if backfill:
                existing = {s["date"] for s in history[code]}
                new_snaps = [s for s in backfill if s["date"] not in existing]
                history[code] = new_snaps + history[code]
                log.info("  %s: added %d backfill points", code, len(new_snaps))

        history[code] = _trim_sort(history[code])
        # Inject downsampled history into markets dict (for HTML)
        markets[code]["history"] = downsample_for_html(history[code])
        log.info("  %s history: %d snapshots for HTML (%d stored)",
                 code, len(markets[code]["history"]), len(history[code]))

    save_history(history)
    # ─────────────────────────────────────────────────────────────────────────

    save_json(markets, timestamp)
    inject_html(markets, timestamp)

    if errors:
        log.warning("Completed with %d error(s):", len(errors))
        for e in errors: log.warning("  * %s", e)
    else:
        log.info("All sources resolved successfully.")

    log.info("Timestamp: %s", timestamp)
    log.info("=" * 60)


if __name__ == "__main__":
    main()


# =============================================================================
#  WINDOWS TASK SCHEDULER SETUP
# =============================================================================
# Run in an elevated Command Prompt to schedule daily at 06:00:
#
#   schtasks /Create /TN "WIRP Daily Update" ^
#     /TR "\"C:\Users\henry\AppData\Local\Programs\Python\Python312\python.exe\" \"C:\Users\henry\OneDrive\Documents\Python\WIRP\update_data.py\"" ^
#     /SC DAILY /ST 06:00 /RU SYSTEM /F
#
# Verify:   schtasks /Query /TN "WIRP Daily Update" /FO LIST
# Test now: schtasks /Run   /TN "WIRP Daily Update"
# Remove:   schtasks /Delete /TN "WIRP Daily Update" /F
#
# Tip: run `where python` to confirm the Python executable path.
# =============================================================================
